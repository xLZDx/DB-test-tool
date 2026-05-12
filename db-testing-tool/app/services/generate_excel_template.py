"""Generate sample Excel template for mapping rules import."""
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

def create_sample_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping Rules"
    
    # Headers
    headers = [
        "Rule Name",
        "Rule Type",
        "Source Datasource ID",
        "Source Schema",
        "Source Table",
        "Source Columns",
        "Target Datasource ID",
        "Target Schema",
        "Target Table",
        "Target Columns",
        "Transformation SQL",
        "Join Condition",
        "Filter Condition",
        "Description"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="002d62", end_color="002d62", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample data row
    sample_data = [
        "Load Customer Data",
        "direct",
        1,
        "CDS_OWNER",
        "CUSTOMERS",
        '["customer_id","customer_name","email"]',
        2,
        "DW_SCHEMA",
        "DIM_CUSTOMER",
        '["cust_id","cust_name","cust_email"]',
        "UPPER(customer_name) AS cust_name",
        "",
        "WHERE active_flag = 'Y'",
        "Loads active customer records from source to target dimension"
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 35
    ws.column_dimensions['K'].width = 40
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 50
    
    # Save
    output_path = Path(__file__).parent.parent.parent / "data" / "mapping_rules_template.xlsx"
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    print(f"Template saved to: {output_path}")

if __name__ == "__main__":
    create_sample_template()

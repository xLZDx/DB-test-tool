"""Excel import service for mapping rules with multiple template support."""
import openpyxl
from typing import List, Dict, Any, Optional, Tuple
import json
import logging
from io import BytesIO

from app.services.template_manager import template_manager, TemplateType, TemplateColumn

logger = logging.getLogger(__name__)


def parse_mapping_rules_excel(file_bytes: bytes, template_type: Optional[TemplateType] = None) -> Dict[str, Any]:
    """
    Parse Excel file containing mapping rules with support for multiple template types.
    
    Args:
        file_bytes: Excel file content as bytes
        template_type: Specific template type to use, or None for auto-detection
    
    Returns: {"rules": [...], "stats": {...}, "errors": [], "detected_template": TemplateType}
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        # Find header row (first non-empty row)
        headers = []
        header_row = 1
        for row in sheet.iter_rows(min_row=1, max_row=10):
            if row[0].value:
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
                break
            header_row += 1
        
        if not headers:
            return {"rules": [], "stats": {}, "errors": ["No header row found in Excel file"], "detected_template": None}
        
        # Auto-detect or validate template type
        detected_template = template_type or template_manager.detect_template_type(headers)
        if not detected_template:
            return {
                "rules": [],
                "stats": {},
                "errors": [f"Unable to detect template type. Found headers: {', '.join(headers)}. Please ensure your Excel file matches one of the supported templates."],
                "detected_template": None
            }
        
        # Get template definition
        template = template_manager.get_template(detected_template)
        
        # Map column headers using template definition
        col_map, mapping_errors = _map_columns_to_template(headers, template)
        
        if mapping_errors:
            return {
                "rules": [],
                "stats": {},
                "errors": mapping_errors,
                "detected_template": detected_template
            }
        
        # Parse data rows
        rules, parsing_errors = _parse_data_rows(sheet, header_row, col_map, template, detected_template)
        
        return {
            "rules": rules,
            "stats": {
                "total_rows": len(rules) + len(parsing_errors),
                "parsed_rules": len(rules),
                "errors": len(parsing_errors),
                "template_type": detected_template.value
            },
            "errors": parsing_errors,
            "detected_template": detected_template
        }
        
    except Exception as e:
        logger.exception("Excel parsing failed")
        return {
            "rules": [], 
            "stats": {}, 
            "errors": [f"Failed to parse Excel file: {str(e)}"],
            "detected_template": None
        }


def _map_columns_to_template(headers: List[str], template) -> Tuple[Dict[str, int], List[str]]:
    """Map Excel headers to template fields."""
    col_map = {}
    errors = []
    
    # Create lookup for headers (case-insensitive)
    header_lookup = {h.lower().strip(): idx for idx, h in enumerate(headers) if h}
    
    # Map each template column
    for col_def in template.columns:
        # Check primary name
        col_names_to_check = [col_def.name.lower()]
        # Add aliases
        col_names_to_check.extend([alias.lower() for alias in col_def.aliases])
        
        # Find matching header
        found = False
        for col_name in col_names_to_check:
            if col_name in header_lookup:
                col_map[col_def.field] = header_lookup[col_name]
                found = True
                break
        
        # Check if required column is missing
        if not found and col_def.required:
            errors.append(f"Missing required column: {col_def.name}. Expected one of: {', '.join([col_def.name] + col_def.aliases)}")
    
    return col_map, errors


def _parse_data_rows(sheet, header_row: int, col_map: Dict[str, int], template, detected_template: TemplateType) -> Tuple[List[Dict], List[str]]:
    """Parse data rows from Excel sheet."""
    rules = []
    errors = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        # Skip empty rows
        if not any(cell.value for cell in row):
            continue
        
        try:
            rule_data = {}
            
            # Extract values based on column mapping
            for field, col_idx in col_map.items():
                value = row[col_idx].value if col_idx < len(row) else None
                
                # Find template column definition for type conversion
                col_def = next((c for c in template.columns if c.field == field), None)
                if not col_def:
                    continue
                
                # Type conversions based on template definition
                if col_def.data_type == "integer":
                    if value:
                        try:
                            rule_data[field] = int(value)
                        except (ValueError, TypeError):
                            errors.append(f"Row {row_idx}: {col_def.name} must be integer, got '{value}'")
                            rule_data[field] = None
                    else:
                        rule_data[field] = None
                elif field in ["source_columns", "target_columns", "column_mappings"]:
                    # Handle JSON arrays or comma-separated values
                    if value:
                        value_str = str(value).strip()
                        if value_str.startswith("[") or value_str.startswith("{"):
                            # Already JSON
                            try:
                                json.loads(value_str)  # Validate JSON
                                rule_data[field] = value_str
                            except json.JSONDecodeError:
                                errors.append(f"Row {row_idx}: {col_def.name} contains invalid JSON: {value_str}")
                                rule_data[field] = None
                        else:
                            # Comma-separated -> JSON array  
                            if "," in value_str:
                                cols = [c.strip() for c in value_str.split(",")]
                                rule_data[field] = json.dumps(cols)
                            else:
                                rule_data[field] = json.dumps([value_str])
                    else:
                        rule_data[field] = None
                else:
                    # String fields
                    rule_data[field] = str(value).strip() if value else None
            
            # Map template-specific fields to standard MappingRule fields
            mapped_rule = _map_template_fields_to_standard(rule_data, detected_template)
            
            # Set default values and validate
            _set_rule_defaults(mapped_rule, detected_template)
            validation_errors = _validate_rule(mapped_rule, template, row_idx)
            
            if validation_errors:
                errors.extend(validation_errors)
                continue
                
            rules.append(mapped_rule)
            
        except Exception as e:
            errors.append(f"Row {row_idx}: Error parsing row - {str(e)}")
            logger.exception(f"Error parsing row {row_idx}")
    
    return rules, errors


def _map_template_fields_to_standard(rule_data: Dict[str, Any], template_type: TemplateType) -> Dict[str, Any]:
    """Map template-specific fields to standard MappingRule fields."""
    mapped_rule = {}
    
    # Standard field mappings that apply to all templates
    standard_fields = [
        "name", "rule_type", "source_datasource_id", "source_schema", "source_table", 
        "source_columns", "target_datasource_id", "target_schema", "target_table", 
        "target_columns", "transformation_sql", "join_condition", "filter_condition", "description"
    ]
    
    # Copy standard fields directly
    for field in standard_fields:
        if field in rule_data:
            mapped_rule[field] = rule_data[field]
    
    # Template-specific field mappings
    if template_type == TemplateType.ENTERPRISE:
        # Map enterprise fields to standard fields
        if not mapped_rule.get("name") and rule_data.get("logical_name"):
            mapped_rule["name"] = rule_data["logical_name"]
        
        if not mapped_rule.get("source_columns") and rule_data.get("source_attribute_name"):
            mapped_rule["source_columns"] = json.dumps([rule_data["source_attribute_name"]])
        
        if not mapped_rule.get("target_columns") and rule_data.get("physical_name"):
            mapped_rule["target_columns"] = json.dumps([rule_data["physical_name"]])
        
        if not mapped_rule.get("transformation_sql") and rule_data.get("transformation_rules"):
            mapped_rule["transformation_sql"] = rule_data["transformation_rules"]
        
        # Enhance description with enterprise metadata
        desc_parts = []
        if rule_data.get("description"):
            desc_parts.append(rule_data["description"])
        if rule_data.get("business_definition"):
            desc_parts.append(f"Business Definition: {rule_data['business_definition']}")
        if rule_data.get("sample_data_values"):
            desc_parts.append(f"Sample Values: {rule_data['sample_data_values']}")
        if rule_data.get("notes_comments"):
            desc_parts.append(f"Notes: {rule_data['notes_comments']}")
        if rule_data.get("pbi_number"):
            desc_parts.append(f"PBI: {rule_data['pbi_number']}")
        
        mapped_rule["description"] = ". ".join(desc_parts) if desc_parts else mapped_rule.get("description")
    
    elif template_type == TemplateType.DATA_WAREHOUSE:
        # Map data warehouse fields
        if not mapped_rule.get("source_datasource_id") and rule_data.get("source_system"):
            # For now, default to datasource 1 - in real implementation, you'd look this up
            mapped_rule["source_datasource_id"] = 1
        
        if not mapped_rule.get("target_datasource_id"):
            # Default to datasource 2 for DWH
            mapped_rule["target_datasource_id"] = 2
        
        # Map dimension/layer info to description
        desc_parts = []
        if rule_data.get("description"):
            desc_parts.append(rule_data["description"])
        if rule_data.get("layer"):
            desc_parts.append(f"Layer: {rule_data['layer']}")
        if rule_data.get("dimension_type"):
            desc_parts.append(f"Type: {rule_data['dimension_type']}")
        if rule_data.get("business_key"):
            desc_parts.append(f"Business Key: {rule_data['business_key']}")
        if rule_data.get("load_strategy"):
            desc_parts.append(f"Load Strategy: {rule_data['load_strategy']}")
        
        mapped_rule["description"] = ". ".join(desc_parts) if desc_parts else mapped_rule.get("description")
    
    elif template_type == TemplateType.ETL_PIPELINE:
        # Map ETL pipeline fields
        if not mapped_rule.get("name") and rule_data.get("pipeline_name"):
            if rule_data.get("step_order"):
                mapped_rule["name"] = f"{rule_data['pipeline_name']} - Step {rule_data['step_order']}"
            else:
                mapped_rule["name"] = rule_data["pipeline_name"]
        
        if not mapped_rule.get("source_table") and rule_data.get("source_object"):
            mapped_rule["source_table"] = rule_data["source_object"]
        
        if not mapped_rule.get("source_schema") and rule_data.get("source_connection"):
            mapped_rule["source_schema"] = rule_data["source_connection"]
        
        if not mapped_rule.get("target_schema") and rule_data.get("target_connection"):
            mapped_rule["target_schema"] = rule_data["target_connection"]
        
        if rule_data.get("source_query"):
            mapped_rule["transformation_sql"] = rule_data["source_query"]
        
        # Build ETL-specific description
        desc_parts = []
        if rule_data.get("description"):
            desc_parts.append(rule_data["description"])
        if rule_data.get("schedule"):
            desc_parts.append(f"Schedule: {rule_data['schedule']}")
        if rule_data.get("dependencies"):
            desc_parts.append(f"Dependencies: {rule_data['dependencies']}")
        if rule_data.get("error_handling"):
            desc_parts.append(f"Error Handling: {rule_data['error_handling']}")
        
        mapped_rule["description"] = ". ".join(desc_parts) if desc_parts else mapped_rule.get("description")
    
    elif template_type == TemplateType.BASIC:
        # Handle simplified basic template - extract schema from table names if present
        # Format: schema.table or just table
        
        # Parse source table for schema
        if mapped_rule.get("source_table") and "." in mapped_rule["source_table"]:
            parts = mapped_rule["source_table"].split(".", 1)
            if not mapped_rule.get("source_schema"):
                mapped_rule["source_schema"] = parts[0]
            mapped_rule["source_table"] = parts[1]
        
        # Parse target table for schema
        if mapped_rule.get("target_table") and "." in mapped_rule["target_table"]:
            parts = mapped_rule["target_table"].split(".", 1)
            if not mapped_rule.get("target_schema"):
                mapped_rule["target_schema"] = parts[0]
            mapped_rule["target_table"] = parts[1]
        
        # Set default datasource IDs if not provided (will use first available datasources)
        # User should edit these after import
        if not mapped_rule.get("source_datasource_id"):
            mapped_rule["source_datasource_id"] = 1  # Default - user should update
        if not mapped_rule.get("target_datasource_id"):
            mapped_rule["target_datasource_id"] = 1  # Default - user should update
    
    return mapped_rule


def _set_rule_defaults(rule: Dict[str, Any], template_type: TemplateType):
    """Set default values based on template type."""
    if not rule.get("rule_type"):
        if template_type == TemplateType.BASIC:
            rule["rule_type"] = "direct"
        elif template_type == TemplateType.ENTERPRISE:
            rule["rule_type"] = "lookup"
        elif template_type == TemplateType.DATA_WAREHOUSE:
            rule["rule_type"] = "scd"
        elif template_type == TemplateType.ETL_PIPELINE:
            rule["rule_type"] = "direct"


def _validate_rule(rule: Dict[str, Any], template, row_idx: int) -> List[str]:
    """Validate parsed rule against MappingRule requirements."""
    errors = []
    
    # Check standard MappingRule required fields 
    required_fields = ["name", "source_table", "target_table"]
    
    for field in required_fields:
        if not rule.get(field):
            errors.append(f"Row {row_idx}: Missing required field: {field}")
    
    # Check datasource IDs are present and valid
    if not rule.get("source_datasource_id"):
        errors.append(f"Row {row_idx}: Missing source datasource ID")
    
    if not rule.get("target_datasource_id"):
        errors.append(f"Row {row_idx}: Missing target datasource ID")
    
    return errors

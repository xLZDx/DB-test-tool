"""Excel template generator service for different mapping template types."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, Any
import json

from app.services.template_manager import template_manager, TemplateType
from app.config import BASE_DIR


class ExcelTemplateGenerator:
    """Generates Excel templates for different mapping scenarios."""
    
    def __init__(self):
        self.template_manager = template_manager
    
    def generate_template(self, template_type: TemplateType, output_path: Path = None) -> Path:
        """Generate Excel template for specified type."""
        template = self.template_manager.get_template(template_type)
        
        if output_path is None:
            output_path = BASE_DIR / "data" / f"mapping_template_{template_type.value}.xlsx"
        
        # Create workbook with template sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = template.name
        
        # Apply header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = []
        for col_idx, col_def in enumerate(template.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_def.name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            headers.append(col_def.name)
            
            # Set column width based on content
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(15, len(col_def.name) + 2)
        
        # Add example data if available
        if template.example_data:
            for row_idx, example_row in enumerate(template.example_data, 2):
                for col_idx, col_def in enumerate(template.columns, 1):
                    if col_def.name in example_row:
                        ws.cell(row=row_idx, column=col_idx).value = example_row[col_def.name]
                    elif col_def.sample_value:
                        ws.cell(row=row_idx, column=col_idx).value = col_def.sample_value
        else:
            # Add one sample row with sample values
            for col_idx, col_def in enumerate(template.columns, 1):
                if col_def.sample_value:
                    ws.cell(row=2, column=col_idx).value = col_def.sample_value
        
        # Create documentation sheet
        doc_ws = wb.create_sheet("Documentation")
        self._create_documentation_sheet(doc_ws, template)
        
        # Create validation sheet for reference data
        if template_type == TemplateType.ENTERPRISE:
            ref_ws = wb.create_sheet("Reference Data")
            self._create_reference_data_sheet(ref_ws)
        
        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    def generate_all_templates(self) -> Dict[TemplateType, Path]:
        """Generate all available template types."""
        paths = {}
        for template_type in TemplateType:
            paths[template_type] = self.generate_template(template_type)
        return paths
    
    def _create_documentation_sheet(self, ws, template):
        """Create documentation sheet with column descriptions."""
        # Title
        ws.cell(row=1, column=1).value = f"{template.name} - Documentation"
        title_font = Font(size=16, bold=True)
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells('A1:D1')
        
        # Description
        ws.cell(row=3, column=1).value = "Description:"
        ws.cell(row=3, column=1).font = Font(bold=True)
        ws.cell(row=3, column=2).value = template.description
        
        # Column documentation header
        ws.cell(row=5, column=1).value = "Column"
        ws.cell(row=5, column=2).value = "Required"
        ws.cell(row=5, column=3).value = "Data Type"
        ws.cell(row=5, column=4).value = "Description"
        ws.cell(row=5, column=5).value = "Sample Value"
        ws.cell(row=5, column=6).value = "Aliases"
        
        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        for col in range(1, 7):
            cell = ws.cell(row=5, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Column documentation
        for row_idx, col_def in enumerate(template.columns, 6):
            ws.cell(row=row_idx, column=1).value = col_def.name
            ws.cell(row=row_idx, column=2).value = "Yes" if col_def.required else "No"
            ws.cell(row=row_idx, column=3).value = col_def.data_type
            ws.cell(row=row_idx, column=4).value = col_def.description
            ws.cell(row=row_idx, column=5).value = col_def.sample_value
            ws.cell(row=row_idx, column=6).value = ", ".join(col_def.aliases) if col_def.aliases else ""
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
    
    def _create_reference_data_sheet(self, ws):
        """Create reference data sheet for enterprise template lookups."""
        ws.cell(row=1, column=1).value = "Reference Data for Enterprise Mapping"
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        # Rule Types
        ws.cell(row=3, column=1).value = "Rule Types:"
        ws.cell(row=3, column=1).font = Font(bold=True)
        rule_types = ["direct", "lookup", "aggregation", "scd", "custom"]
        for i, rt in enumerate(rule_types, 4):
            ws.cell(row=i, column=1).value = rt
        
        # Actions
        ws.cell(row=3, column=3).value = "Actions:"
        ws.cell(row=3, column=3).font = Font(bold=True)
        actions = ["Add", "Update", "Remove", "No Change"]
        for i, action in enumerate(actions, 4):
            ws.cell(row=i, column=3).value = action
        
        # Data Types (Oracle)
        ws.cell(row=3, column=5).value = "Oracle Data Types:"
        ws.cell(row=3, column=5).font = Font(bold=True)
        oracle_types = ["VARCHAR2(n)", "NUMBER(m,n)", "DATE", "TIMESTAMP(6)", "CLOB", "BLOB"]
        for i, ot in enumerate(oracle_types, 4):
            ws.cell(row=i, column=5).value = ot
        
        # Data Types (Redshift)
        ws.cell(row=3, column=7).value = "Redshift Data Types:"
        ws.cell(row=3, column=7).font = Font(bold=True)
        redshift_types = ["character varying(n)", "numeric(m,n)", "date", "timestamp", "text", "boolean"]
        for i, rt in enumerate(redshift_types, 4):
            ws.cell(row=i, column=7).value = rt
        
        # Sample transformations
        ws.cell(row=12, column=1).value = "Sample Transformations:"
        ws.cell(row=12, column=1).font = Font(bold=True)
        transformations = [
            "Direct mapping: source_col",
            "Lookup: LEFT JOIN lookup_table ON source.key = lookup.key",
            "Conversion: UPPER(TRIM(source_col))",
            "Aggregation: SUM(amount) GROUP BY customer_id",
            "Date format: TO_DATE(source_date, 'YYYY-MM-DD')"
        ]
        for i, trans in enumerate(transformations, 13):
            ws.cell(row=i, column=1).value = trans
        
        # Adjust column widths
        for col in ['A', 'C', 'E', 'G']:
            ws.column_dimensions[col].width = 25


def create_sample_templates():
    """Create sample templates for all types."""
    generator = ExcelTemplateGenerator()
    paths = generator.generate_all_templates()
    
    print("Generated Excel templates:")
    for template_type, path in paths.items():
        print(f"  {template_type.value}: {path}")
    
    return paths


if __name__ == "__main__":
    create_sample_templates()